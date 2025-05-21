import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from Plantilla import ajustar_ancho
import shutil

# Carpeta de descargas por defecto
ruta_carpeta_descargas = os.path.join(os.path.expanduser("~"), "Downloads")

def generar_resultados(datos, resultados_df, nombre_archivo_salida="resultados_certificados.xlsx", carpeta_destino=ruta_carpeta_descargas):
    # Cambia 'resultados' por 'resultados_df'
    if isinstance(resultados_df, list):
        resultados_df = pd.DataFrame(resultados_df)

    if resultados_df.empty:
        print("Error: El DataFrame 'resultados' está vacío. No se puede generar el archivo.")
        return

    # Carpeta de descargas
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    if not os.path.exists(downloads_folder):
        print("Error: No se puede encontrar la carpeta de Descargas.")
        return

    # Ruta general de salida (ya no por ficha)
    ruta_archivo_salida = os.path.join(downloads_folder, nombre_archivo_salida)

    # Agregar columna de observaciones directamente al DataFrame completo
    print("Agregando columna de observaciones...")
    datos["OBSERVACIONES"] = resultados_df["OBSERVACIONES"]  # Cambia 'resultados' por 'resultados_df'

    # Guardar el archivo
    try:
        datos.to_excel(ruta_archivo_salida, index=False, engine="openpyxl")
        print(f"Archivo guardado en: {ruta_archivo_salida}")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")
        return

    # Ajustar ancho de columnas
    try:
        ajustar_ancho(ruta_archivo_salida)
    except Exception as e:
        print(f"Error al ajustar el ancho de las columnas: {e}")
        return

    # Aplicar colores
    try:
        wb = load_workbook(ruta_archivo_salida)
        ws = wb.active

        for idx, row in resultados_df.iterrows():  # Cambia 'resultados' por 'resultados_df'
            fill_color = None

            if pd.isna(row.get("STATUS")) or pd.isna(row.get("OBSERVACIONES")):
                continue

            if row["STATUS"] == "EXITO":
                fill_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif row["STATUS"] == "NOVEDAD":
                fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif row["STATUS"] == "FALLIDO":
                fill_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            if fill_color:
                excel_row = idx + 2  # Encabezado
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=excel_row, column=col).fill = fill_color

        wb.save(ruta_archivo_salida)
        print(f"Archivo coloreado y guardado en: {ruta_archivo_salida}")
    except Exception as e:
        print(f"Error al aplicar colores: {e}")
        return

    # Mover si hay carpeta destino
    if carpeta_destino:
        try:
            shutil.move(ruta_archivo_salida, os.path.join(carpeta_destino, nombre_archivo_salida))
            print(f"Archivo movido a: {os.path.join(carpeta_destino, nombre_archivo_salida)}")
        except Exception as e:
            print(f"Error al mover el archivo a la carpeta de destino: {e}")
